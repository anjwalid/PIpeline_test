from __future__ import annotations

from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from app.repositories.catalog_repository import CatalogRepository


class CatalogExportService:
    TEMPLATE_PATH = (
        Path(__file__).resolve().parents[2]
        / "resources"
        / "templates"
        / "Template_excel.xlsx"
    )
    EXPORT_HEADERS = (
        "ID menace",
        "Menace",
        "Description",
        "Scenarios d'attaque",
        "Mesures de securite recommandees",
        "References",
    )
    DATA_START_ROW = 10

    @classmethod
    def build_workbook_bytes(cls) -> bytes:
        workbook = load_workbook(cls.TEMPLATE_PATH)
        worksheet = workbook[workbook.sheetnames[0]]

        style_row = cls._capture_style_row(worksheet, cls.DATA_START_ROW)
        cls._prepare_worksheet(worksheet)

        threats = CatalogRepository.list_threats_for_analysis()
        current_row = cls.DATA_START_ROW

        for threat in threats:
            scenarios = cls._normalize_items(
                item.get("description_scenario") for item in threat.get("scenarios", [])
            )
            mitigations = cls._normalize_items(
                item.get("description_mitigation")
                for item in threat.get("mitigations", [])
            )
            references = cls._normalize_items(
                cls._format_reference(item) for item in threat.get("references", [])
            )

            span = max(len(scenarios), len(mitigations), len(references), 1)
            start_row = current_row
            end_row = current_row + span - 1

            cls._write_group_value(
                worksheet, start_row, end_row, 1, threat.get("id_menace"), style_row[0]
            )
            cls._write_group_value(
                worksheet, start_row, end_row, 2, threat.get("nom_menace"), style_row[1]
            )
            cls._write_group_value(
                worksheet, start_row, end_row, 3, threat.get("description"), style_row[2]
            )

            for offset in range(span):
                row_index = start_row + offset
                worksheet.row_dimensions[row_index].height = 24
                cls._write_cell(
                    worksheet,
                    row_index,
                    4,
                    scenarios[offset] if offset < len(scenarios) else "",
                    style_row[3],
                )
                cls._write_cell(
                    worksheet,
                    row_index,
                    5,
                    mitigations[offset] if offset < len(mitigations) else "",
                    style_row[4],
                )
                cls._write_cell(
                    worksheet,
                    row_index,
                    6,
                    references[offset] if offset < len(references) else "",
                    style_row[5],
                )

            current_row = end_row + 1

        final_row = max(current_row - 1, cls.DATA_START_ROW)
        worksheet.auto_filter.ref = f"A9:F{final_row}"
        worksheet.freeze_panes = None

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    def build_filename(cls) -> str:
        timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
        return f"catalogue-menaces-awb-{timestamp}.xlsx"

    @classmethod
    def _prepare_worksheet(cls, worksheet: Worksheet) -> None:
        worksheet.column_dimensions["F"].hidden = False
        worksheet.column_dimensions["F"].width = 42

        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.min_row >= cls.DATA_START_ROW:
                worksheet.unmerge_cells(str(merged_range))

        if worksheet.max_row >= cls.DATA_START_ROW:
            worksheet.delete_rows(cls.DATA_START_ROW, worksheet.max_row - cls.DATA_START_ROW + 1)

        for column_index, header in enumerate(cls.EXPORT_HEADERS, start=1):
            worksheet.cell(row=9, column=column_index).value = header

    @staticmethod
    def _capture_style_row(worksheet: Worksheet, row_index: int):
        return [copy(worksheet.cell(row=row_index, column=column)) for column in range(1, 7)]

    @classmethod
    def _write_group_value(
        cls,
        worksheet: Worksheet,
        start_row: int,
        end_row: int,
        column_index: int,
        value: object,
        template_cell,
    ) -> None:
        cls._write_cell(worksheet, start_row, column_index, value, template_cell)

        if end_row > start_row:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=column_index,
                end_row=end_row,
                end_column=column_index,
            )

    @staticmethod
    def _write_cell(
        worksheet: Worksheet,
        row_index: int,
        column_index: int,
        value: object,
        template_cell,
    ) -> None:
        cell = worksheet.cell(row=row_index, column=column_index)
        cell.value = CatalogExportService._sanitize_excel_value(value)
        cell._style = copy(template_cell._style)
        cell.font = copy(template_cell.font)
        cell.fill = copy(template_cell.fill)
        cell.border = copy(template_cell.border)
        cell.protection = copy(template_cell.protection)
        cell.number_format = template_cell.number_format
        existing_alignment = copy(template_cell.alignment)
        cell.alignment = Alignment(
            horizontal=(existing_alignment.horizontal or "left"),
            vertical="top",
            wrap_text=True,
        )

    @staticmethod
    def _sanitize_excel_value(value: object) -> object:
        if not isinstance(value, str):
            return value

        normalized = value.strip()
        if not normalized:
            return ""

        if normalized[0] in ("=", "+", "-", "@"):
            return f"'{normalized}"

        return normalized

    @staticmethod
    def _normalize_items(values: Iterable[object]) -> list[str]:
        normalized = []
        for value in values:
            sanitized = CatalogExportService._sanitize_excel_value(value)
            if isinstance(sanitized, str) and sanitized:
                normalized.append(sanitized)
        return normalized

    @staticmethod
    def _format_reference(reference: dict) -> str:
        code = (reference.get("reference_menace") or "").strip()
        label = (reference.get("nom_reference") or "").strip()

        if code and label:
            return f"{code} - {label}"
        return code or label
